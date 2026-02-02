"""
Export Module - Phase 3
Bangko ng Seton

Handles data export to various formats (CSV, Excel)
"""
import csv
import io
from datetime import datetime, timedelta
from typing import List, Dict, Optional
import pytz

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

PHILIPPINES_TZ = pytz.timezone('Asia/Manila')


def get_philippines_time():
    """Get current time in Philippine timezone"""
    return datetime.now(PHILIPPINES_TZ)


class DataExporter:
    """Handles data export to various formats"""
    
    def __init__(self, data: List[Dict], export_type: str = 'transactions'):
        """
        Initialize exporter with data
        
        Args:
            data: List of dictionaries to export
            export_type: Type of data ('transactions', 'students', 'accounts')
        """
        self.data = data
        self.export_type = export_type
    
    def to_csv(self) -> str:
        """
        Export data to CSV format
        
        Returns:
            CSV string
        """
        if not self.data:
            return ""
        
        output = io.StringIO()
        
        # Get headers from first record
        headers = list(self.data[0].keys())
        
        # Remove internal fields
        headers = [h for h in headers if not h.startswith('_')]
        
        writer = csv.DictWriter(output, fieldnames=headers, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(self.data)
        
        return output.getvalue()
    
    def to_excel(self) -> Optional[bytes]:
        """
        Export data to Excel format
        
        Returns:
            Excel file as bytes, or None if openpyxl not available
        """
        if not EXCEL_AVAILABLE:
            return None
        
        if not self.data:
            return None
        
        wb = Workbook()
        ws = wb.active
        ws.title = self.export_type.capitalize()
        
        # Get headers
        headers = [h for h in self.data[0].keys() if not h.startswith('_')]
        
        # Style headers
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, record in enumerate(self.data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = record.get(header, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, max_row=len(self.data) + 1, 
                                   min_col=col_idx, max_col=col_idx):
                cell_value = str(row[0].value) if row[0].value else ''
                max_length = max(max_length, len(cell_value))
            
            # Set column width (max 50 characters)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()


def filter_by_date_range(data: List[Dict], 
                        start_date: Optional[str] = None,
                        end_date: Optional[str] = None,
                        date_field: str = 'Date') -> List[Dict]:
    """
    Filter data by date range
    
    Args:
        data: List of dictionaries with date field
        start_date: Start date (YYYY-MM-DD format)
        end_date: End date (YYYY-MM-DD format)
        date_field: Name of date field in data
    
    Returns:
        Filtered list of dictionaries
    """
    if not start_date and not end_date:
        return data
    
    filtered = []
    
    # Parse date range
    start_dt = None
    end_dt = None
    
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            start_dt = PHILIPPINES_TZ.localize(start_dt)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            end_dt = PHILIPPINES_TZ.localize(end_dt.replace(hour=23, minute=59, second=59))
        except ValueError:
            pass
    
    for record in data:
        record_date_str = record.get(date_field, '')
        
        if not record_date_str:
            continue
        
        # Parse record date
        try:
            # Try ISO format
            record_date = datetime.fromisoformat(record_date_str.replace('Z', '+00:00'))
            # Ensure timezone aware
            if record_date.tzinfo is None:
                record_date = PHILIPPINES_TZ.localize(record_date)
        except:
            try:
                # Try common format
                record_date = datetime.strptime(record_date_str, '%Y-%m-%d %H:%M:%S')
                record_date = PHILIPPINES_TZ.localize(record_date)
            except:
                continue
        
        # Apply filters
        if start_dt and record_date < start_dt:
            continue
        if end_dt and record_date > end_dt:
            continue
        
        filtered.append(record)
    
    return filtered


def export_transactions(transactions: List[Dict],
                       format: str = 'csv',
                       start_date: Optional[str] = None,
                       end_date: Optional[str] = None) -> tuple:
    """
    Export transactions with date filtering
    
    Args:
        transactions: List of transaction dictionaries
        format: Export format ('csv' or 'excel')
        start_date: Start date filter (YYYY-MM-DD)
        end_date: End date filter (YYYY-MM-DD)
    
    Returns:
        Tuple of (data, mimetype, filename)
    """
    # Filter by date
    filtered = filter_by_date_range(transactions, start_date, end_date)
    
    # Generate filename
    now = get_philippines_time()
    date_range = ''
    if start_date and end_date:
        date_range = f"_{start_date}_to_{end_date}"
    elif start_date:
        date_range = f"_from_{start_date}"
    elif end_date:
        date_range = f"_until_{end_date}"
    
    timestamp = now.strftime('%Y%m%d_%H%M%S')
    
    exporter = DataExporter(filtered, 'transactions')
    
    if format.lower() == 'excel' and EXCEL_AVAILABLE:
        data = exporter.to_excel()
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = f'transactions{date_range}_{timestamp}.xlsx'
    else:
        # Default to CSV
        data = exporter.to_csv()
        mimetype = 'text/csv'
        filename = f'transactions{date_range}_{timestamp}.csv'
    
    return data, mimetype, filename


def export_students(students: List[Dict], format: str = 'csv') -> tuple:
    """
    Export student list
    
    Args:
        students: List of student dictionaries
        format: Export format ('csv' or 'excel')
    
    Returns:
        Tuple of (data, mimetype, filename)
    """
    now = get_philippines_time()
    timestamp = now.strftime('%Y%m%d_%H%M%S')
    
    exporter = DataExporter(students, 'students')
    
    if format.lower() == 'excel' and EXCEL_AVAILABLE:
        data = exporter.to_excel()
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = f'students_{timestamp}.xlsx'
    else:
        data = exporter.to_csv()
        mimetype = 'text/csv'
        filename = f'students_{timestamp}.csv'
    
    return data, mimetype, filename


def generate_monthly_statement(student_id: str,
                               transactions: List[Dict],
                               student_name: str,
                               month: Optional[str] = None) -> str:
    """
    Generate a text-based monthly statement for a student
    
    Args:
        student_id: Student ID
        transactions: All transactions
        student_name: Student name
        month: Month in YYYY-MM format (defaults to current month)
    
    Returns:
        Formatted statement as string
    """
    now = get_philippines_time()
    
    if month:
        try:
            year, month_num = map(int, month.split('-'))
            start_date = datetime(year, month_num, 1)
            # Get last day of month
            if month_num == 12:
                end_date = datetime(year + 1, 1, 1) - timedelta(seconds=1)
            else:
                end_date = datetime(year, month_num + 1, 1) - timedelta(seconds=1)
            start_date = PHILIPPINES_TZ.localize(start_date)
            end_date = PHILIPPINES_TZ.localize(end_date)
        except:
            # Default to current month
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if now.month == 12:
                end_date = now.replace(year=now.year + 1, month=1, day=1)
            else:
                end_date = now.replace(month=now.month + 1, day=1)
            end_date = end_date - timedelta(seconds=1)
    else:
        start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if now.month == 12:
            end_date = now.replace(year=now.year + 1, month=1, day=1)
        else:
            end_date = now.replace(month=now.month + 1, day=1)
        end_date = end_date - timedelta(seconds=1)
    
    # Filter transactions for this student and date range
    student_txns = []
    for txn in transactions:
        if txn.get('StudentID') != student_id:
            continue
        
        txn_date_str = txn.get('Date', '')
        try:
            txn_date = datetime.fromisoformat(txn_date_str.replace('Z', '+00:00'))
        except:
            try:
                txn_date = datetime.strptime(txn_date_str, '%Y-%m-%d %H:%M:%S')
                txn_date = PHILIPPINES_TZ.localize(txn_date)
            except:
                continue
        
        if start_date <= txn_date <= end_date:
            student_txns.append({
                'date': txn_date,
                'type': txn.get('Type', ''),
                'amount': float(txn.get('Amount', 0))
            })
    
    # Sort by date
    student_txns.sort(key=lambda x: x['date'])
    
    # Generate statement
    statement = []
    statement.append("=" * 60)
    statement.append("BANGKO NG SETON - MONTHLY STATEMENT")
    statement.append("=" * 60)
    statement.append(f"\nStudent: {student_name}")
    statement.append(f"Student ID: {student_id}")
    statement.append(f"Statement Period: {start_date.strftime('%B %Y')}")
    statement.append(f"Generated: {now.strftime('%Y-%m-%d %H:%M:%S')}")
    statement.append("\n" + "-" * 60)
    statement.append(f"{'Date':<20} {'Type':<15} {'Amount':>10}")
    statement.append("-" * 60)
    
    total_loaded = 0.0
    total_spent = 0.0
    
    for txn in student_txns:
        date_str = txn['date'].strftime('%Y-%m-%d %H:%M')
        txn_type = txn['type']
        amount = txn['amount']
        
        if txn_type == 'Load':
            amount_str = f"+₱{amount:.2f}"
            total_loaded += amount
        else:
            amount_str = f"-₱{abs(amount):.2f}"
            total_spent += abs(amount)
        
        statement.append(f"{date_str:<20} {txn_type:<15} {amount_str:>10}")
    
    statement.append("-" * 60)
    statement.append(f"{'Total Loaded:':<35} ₱{total_loaded:.2f}")
    statement.append(f"{'Total Spent:':<35} ₱{total_spent:.2f}")
    statement.append(f"{'Net Change:':<35} ₱{total_loaded - total_spent:.2f}")
    statement.append("=" * 60)
    statement.append("\nThank you for using Bangko ng Seton!")
    statement.append("For questions, contact the Finance Office.")
    statement.append("=" * 60)
    
    return "\n".join(statement)
